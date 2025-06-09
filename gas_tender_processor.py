# Generic Gas Tender Processing System
# Combines functionality from create_csv, extract_mibgas, add_columns, and prop_dgn scripts
# with company-specific references removed for confidentiality

from openpyxl import load_workbook
import csv
import re
from datetime import datetime, timedelta, date
import calendar
import pandas as pd
import argparse
from pathlib import Path


class GasTenderProcessor:
    """
    Generic Gas Tender Processing System
    Handles CSV creation, MIBGAS data extraction, column additions, and proposal calculations
    """
    
    def __init__(self):
        # Standard headers for gas tender CSV files
        self.base_headers = [
            'N.º CONCURSO', 'REFERÊNCIA', 'TIPO', 'NOME', 'DATA DA PUBLICAÇÃO',
            'PRAZO DE ENTREGA.DATA', 'PRAZO DE ENTREGA.HORA',
            'Pedido de Esclarecimentos', 'Plataforma', 'CONSUMO TOTAL.kWh',
            'CONSUMO TOTAL.m3', 'ESCALÕES.< 10.000', 'ESCALÕES.> 10.000',
            'PRAZOS CONTRATUAIS.DE FORNECIMENTO', 'PRAZOS CONTRATUAIS.INICIO',
            'PRAZOS CONTRATUAIS.FIM', 'VALOR CONTRATO',
            'PROPOSTO PRINCIPAL (€/kWh).Indexante.TTF/Mib/Outro',
            'PROPOSTO PRINCIPAL (€/kWh).Indexante.Cotação', 
            'PROPOSTO PRINCIPAL (€/kWh).K',
            'PROPOSTO PRINCIPAL (€/kWh).PE',
            'PROPOSTA CONCORRENTES €/kWh.Empresa1',
            'PROPOSTA CONCORRENTES €/kWh.Empresa2',
            'PROPOSTA CONCORRENTES €/kWh.Empresa3',
            'PROPOSTA CONCORRENTES €/kWh.Empresa4',
            'PROPOSTA CONCORRENTES €/kWh.Empresa5',
            'PROPOSTA CONCORRENTES €/kWh.Outros',
            'VENCEDOR', 'DATA DE CONCLUSÃO'
        ]
    
    def clean_value(self, value):
        """Clean values by removing special characters and normalizing format"""
        if value is None or (isinstance(value, float) and value != value):
            return ''
        if isinstance(value, str):
            value = re.sub(r'[\n\r\t"\';]', ' ', value.strip())
            value = re.sub(r'\s+', ' ', value)
            value = re.sub(r'(?<=\d)\.(?=\d)', ',', value)
            if value.lower() in ['none', 'nan', '']:
                return ''
        return str(value)

    def format_date(self, value):
        """Format dates to DD/MM/YYYY format"""
        try:
            if isinstance(value, (datetime, date)):
                return value.strftime('%d/%m/%Y')
            if isinstance(value, (int, float)):
                return (datetime(1899, 12, 30) + timedelta(days=value)).strftime('%d/%m/%Y')
            if isinstance(value, str):
                for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y'):
                    try:
                        return datetime.strptime(value, fmt).strftime('%d/%m/%Y')
                    except:
                        continue
        except:
            pass
        return '-'

    def clean_number(self, value):
        """Clean and format numeric values"""
        if isinstance(value, str) and value.strip().upper() == "NAP":
            return "NAP"
        
        cleaned = self.clean_value(value)
        if cleaned == '-': 
            return '-'
        
        cleaned = re.sub(r'[^\d,.-]', '', cleaned)
        if cleaned.count(',') > 1 or cleaned.count('.') > 1:
            parts = re.split(r'[,.]', cleaned)
            if len(parts) > 2:
                cleaned = f"{''.join(parts[:-1])},{parts[-1]}"
        return cleaned.replace('.', ',').replace(' ', '')

    def clean_multiline_value(self, value):
        """Clean values that might contain multiple lines or entries"""
        if isinstance(value, str):
            value = re.sub(r'\s+', ' ', value)
            value = re.sub(r'[";]+$', '', value)
            value = value.replace('"', '')
            
            if 'TTF' in value and value.count('TTF') > 1:
                parts = value.split('TTF')
                return f"TTF{parts[1].split(';')[0]}"
                
            if re.search(r'\d+[,\.]\d+.*\d+[,\.]\d+', value):
                numbers = re.findall(r'\d+[,\.]\d+', value)
                if numbers:
                    return numbers[0].replace('.', ',')
            
            value = re.sub(r'\s+', ' ', value)
        return value

    def is_empty_row(self, row_dict):
        """Check if a row is effectively empty"""
        empty_count = sum(1 for value in row_dict.values() if not value or value == '-' or value == 'None')
        return empty_count >= len(row_dict.values()) - 2

    def parse_date(self, date_str):
        """Parse date string in DD/MM/YYYY format"""
        if date_str and date_str != '-':
            try:
                return datetime.strptime(date_str, '%d/%m/%Y')
            except ValueError:
                return None
        return None

    def get_month_name(self, month_num):
        """Get month name from number"""
        return calendar.month_name[month_num]

    def calculate_supply_month(self, price_date, inicio_value):
        """Calculate start supply month based on price date and contract start"""
        if inicio_value != 'Ass.' and inicio_value != '-':
            inicio_date = self.parse_date(inicio_value)
            if inicio_date:
                return f"{inicio_date.month:02d}-{self.get_month_name(inicio_date.month)}"
            return '-'
        
        if not price_date:
            return '-'
        
        day = price_date.day
        month = price_date.month
        year = price_date.year
        
        months_to_add = 2 if day < 21 else 3
        
        new_month = month + months_to_add
        new_year = year
        
        if new_month > 12:
            new_month -= 12
            new_year += 1
        
        month_name = self.get_month_name(new_month)
        return f"{new_month:02d}-{month_name}"

    def process_header_sheet(self, sheet):
        """Extract headers from Excel sheet and handle merged cells"""
        merged_ranges = sheet.merged_cells.ranges
        headers = []
        for row_idx in [1, 2, 3]:
            row = [cell.value for cell in sheet[row_idx]]
            for merge in merged_ranges:
                if merge.min_row == row_idx and merge.max_row == row_idx:
                    top_left_value = sheet.cell(row=row_idx, column=merge.min_col).value
                    for col in range(merge.min_col, merge.max_col + 1):
                        row[col - 1] = top_left_value
            headers.append(row)
        columns = list(zip(*headers))
        return [tuple(col) for col in columns]

    def generate_unique_labels(self, header_tuples):
        """Generate clean and unique column labels"""
        labels = []
        seen = {}
        for parts in header_tuples:
            cleaned = [re.sub(r'\s+', ' ', str(p).strip()) if p else '' for p in parts]
            label = '.'.join(filter(None, cleaned))
            if not label:
                continue
            if label in seen:
                seen[label] += 1
                label = f"{label}_{seen[label]}"
            else:
                seen[label] = 0
            labels.append(label)
        return labels

    def create_csv_from_excel(self, excel_path, target_sheet_name, data_sheet_name, output_csv, year=None):
        """Create CSV from Excel file with proper data cleaning and formatting"""
        wb = load_workbook(excel_path, data_only=True)
        
        if year and year >= 2024:
            # Use newer processing method for 2024+
            target_sheet = wb[target_sheet_name]
            target_tuples = self.process_header_sheet(target_sheet)
            target_labels = self.generate_unique_labels(target_tuples)
            
            data_sheet = wb[data_sheet_name]
            data_tuples = self.process_header_sheet(data_sheet)
            data_header_indices = {header: idx for idx, header in enumerate(data_tuples)}
            
            target_indices = [data_header_indices.get(header, -1) for header in target_tuples]
            
            data_rows = []
            max_row = 92 if year == 2024 else 105
            for row in data_sheet.iter_rows(min_row=4, max_row=max_row, values_only=True):
                formatted_row = []
                for ti in target_indices:
                    value = row[ti] if ti != -1 and ti < len(row) else '-'
                    formatted_value = self.format_date(value) if isinstance(value, (datetime, date)) else value
                    formatted_row.append(self.clean_value(str(formatted_value)))
                
                formatted_row = formatted_row[:len(target_labels)]
                data_rows.append(formatted_row)
            
            with open(output_csv, 'w', newline='', encoding='utf-8-sig') as f:
                f.write(';'.join(target_labels) + '\n')
                for row in data_rows:
                    f.write(';'.join(str(cell) for cell in row) + '\n')
        else:
            # Use legacy processing for older years
            sheet = wb.active
            
            with open(output_csv, 'w', newline='', encoding='utf-8-sig') as csvfile:
                writer = csv.DictWriter(csvfile, fieldnames=self.base_headers, delimiter=';')
                writer.writeheader()
                
                for row in sheet.iter_rows(min_row=4, max_row=105, values_only=True):
                    processed = {
                        'N.º CONCURSO': str(row[0]),
                        'REFERÊNCIA': str(row[1]),
                        'TIPO': str(row[2]),
                        'NOME': str(row[3]).strip(),
                        'DATA DA PUBLICAÇÃO': self.format_date(row[4]),
                        'PRAZO DE ENTREGA.DATA': self.format_date(row[5]),
                        'PRAZO DE ENTREGA.HORA': str(row[6]),
                        'Pedido de Esclarecimentos': '',
                        'Plataforma': '',
                        'CONSUMO TOTAL.kWh': self.clean_number(str(row[7])),
                        'CONSUMO TOTAL.m3': self.clean_number(str(row[8])),
                        'ESCALÕES.< 10.000': self.clean_number(str(row[9])),
                        'ESCALÕES.> 10.000': self.clean_number(str(row[10])),
                        'PRAZOS CONTRATUAIS.DE FORNECIMENTO': self.clean_multiline_value(str(row[11])),
                        'PRAZOS CONTRATUAIS.INICIO': self.format_date(row[12]),
                        'PRAZOS CONTRATUAIS.FIM': self.format_date(row[13]),
                        'VALOR CONTRATO': self.clean_number(str(row[14]).replace(' €', '')),
                        'PROPOSTO PRINCIPAL (€/kWh).Indexante.TTF/Mib/Outro': self.clean_multiline_value(str(row[15])),
                        'PROPOSTO PRINCIPAL (€/kWh).Indexante.Cotação': self.clean_multiline_value(self.clean_number(str(row[16]))),
                        'PROPOSTO PRINCIPAL (€/kWh).K': self.clean_multiline_value(self.clean_number(str(row[17]))),
                        'PROPOSTO PRINCIPAL (€/kWh).PE': str(row[18]).strip() if str(row[18]).strip().upper() == "NAP" else self.clean_multiline_value(self.clean_number(str(row[18]))),
                        'PROPOSTA CONCORRENTES €/kWh.Empresa1': self.clean_multiline_value(str(row[19])),
                        'PROPOSTA CONCORRENTES €/kWh.Empresa2': self.clean_multiline_value(str(row[20])),
                        'PROPOSTA CONCORRENTES €/kWh.Empresa3': self.clean_multiline_value(str(row[21])),
                        'PROPOSTA CONCORRENTES €/kWh.Empresa4': self.clean_multiline_value(str(row[22])),
                        'PROPOSTA CONCORRENTES €/kWh.Empresa5': self.clean_multiline_value(str(row[23])),
                        'PROPOSTA CONCORRENTES €/kWh.Outros': self.clean_multiline_value(str(row[24])),
                        'VENCEDOR': self.clean_value(row[25] if len(row) > 25 else ''),
                        'DATA DE CONCLUSÃO': self.format_date(row[26] if len(row) > 26 else '')
                    }
                    
                    for key in processed:
                        processed[key] = (
                            str(processed[key])
                            .replace('..', ',')
                            .replace('.,', ',')
                            .replace(',,', ',')
                            .replace('-.', '-')
                            .replace('None', '')
                        )
                    
                    if not self.is_empty_row(processed):
                        writer.writerow(processed)

    def extract_mibgas_data(self, mibgas_excel, csv_file):
        """Create columns of the proposal of the winners and the competitors"""
        # Read existing CSV
        with open(csv_file, 'r', encoding='utf-8-sig') as f:
            content = f.read().strip()
            delimiter = ';' if ';' in content.split('\n')[0] else ','
            f.seek(0)
            
            csv_reader = csv.reader(f, delimiter=delimiter, quotechar='"')
            headers = next(csv_reader)
            rows = list(csv_reader)
        
        # Find required column indices
        required_columns = {
            'date': "PRAZO DE ENTREGA.DATA",
            'winner': "VENCEDOR", 
            'k': "PROPOSTO PRINCIPAL (€/kWh).K",
            'pe': "PROPOSTO PRINCIPAL (€/kWh).PE",
            'inicio': "PRAZOS CONTRATUAIS.INICIO",
            'indexante': "PROPOSTO PRINCIPAL (€/kWh).Indexante.TTF/Mib/Outro"
        }
        
        col_indices = {}
        for key, col_name in required_columns.items():
            for idx, header in enumerate(headers):
                if header.strip() == col_name:
                    col_indices[key] = idx
                    break
        
        # Process data
        winner_proposals = []
        companies_with_proposals = []
        price_dates = []
        start_supply_months = []
        
        for row_idx, row in enumerate(rows, start=1):
            try:
                if len(row) <= max(col_indices.values()):
                    winner_proposals.append('-')
                    companies_with_proposals.append('-')
                    price_dates.append('-')
                    start_supply_months.append('-')
                    continue
                
                date_str = row[col_indices['date']].strip()
                winner = row[col_indices['winner']].strip()
                inicio_value = row[col_indices['inicio']].strip()
                indexante = row[col_indices['indexante']].strip()
                
                # Calculate Price Date
                prazo_data = self.parse_date(date_str)
                price_date = None
                if prazo_data:
                    price_date = prazo_data - timedelta(days=2)
                    if indexante == "Fixo":
                        price_dates.append(self.format_date(price_date))
                    else:
                        price_dates.append('-')
                else:
                    price_dates.append('-')
                
                # Calculate Start supply month
                start_supply_months.append(self.calculate_supply_month(price_date, inicio_value))
                
                # Find companies with proposals
                companies = []
                for header in headers:
                    if 'PROPOSTA CONCORRENTES €/kWh.' in header:
                        company = header.replace('PROPOSTA CONCORRENTES €/kWh.', '').split(' ')[0]
                        col_idx = headers.index(header)
                        if len(row) > col_idx:
                            value = row[col_idx].strip()
                            if value and value not in ['-', 'NAP', 'None', '']:
                                if company not in companies:
                                    companies.append(company)
                
                # Check main company proposals
                pe_value = row[col_indices['pe']].strip()
                k_value = row[col_indices['k']].strip()
                
                if ((pe_value and pe_value not in ['-', 'NAP', 'None', '']) or 
                    (k_value and k_value not in ['-', 'NAP', 'None', ''])):
                    if 'PRINCIPAL' not in companies:
                        companies.append('PRINCIPAL')
                
                companies_with_proposals.append(', '.join(sorted(companies)) if companies else '-')
                
                # Process winner proposal
                proposal_value = self._process_winner_proposal(row, headers, winner, col_indices, indexante)
                winner_proposals.append(proposal_value)
                
            except Exception as e:
                print(f"Row {row_idx}: Error processing row: {e}")
                winner_proposals.append('-')
                companies_with_proposals.append('-')
                price_dates.append('-')
                start_supply_months.append('-')
        
        # Update CSV with new data
        self._write_updated_csv(csv_file, headers, rows, price_dates, start_supply_months, 
                               winner_proposals, companies_with_proposals, delimiter)

    def _process_winner_proposal(self, row, headers, winner, col_indices, indexante):
        """Process winner proposal value"""
        proposal_value = None
        
        if winner and winner not in ['-', 'todos', 'todos excluídos', 'cancelado', 'em espera']:
            if winner == "PRINCIPAL":
                pe_value = row[col_indices['pe']].strip()
                k_value = row[col_indices['k']].strip()
                
                if indexante and indexante in ["TTF (1 -1, 1)", "MibGás (1 -1, 1)"] and k_value and k_value != '-':
                    try:
                        k_float = float(k_value.replace(',', '.'))
                        mwh_value = k_float * 1000
                        formatted_value = f"{mwh_value:.3f}".replace('.', ',')
                        return f"k={formatted_value}"
                    except (ValueError, AttributeError):
                        pass
                
                if pe_value and pe_value != '-':
                    proposal_value = pe_value
                elif k_value and k_value != '-':
                    proposal_value = k_value
            else:
                # Look for competitor proposal
                possible_headers = [
                    f"PROPOSTA CONCORRENTES €/kWh.{winner}",
                    f"PROPOSTA CONCORRENTES €/kWh.{winner} Energy"
                ]
                
                for proposal_header in possible_headers:
                    try:
                        proposal_col_idx = headers.index(proposal_header.strip())
                        if len(row) > proposal_col_idx:
                            proposal_value = row[proposal_col_idx].strip()
                            if proposal_value:
                                break
                    except ValueError:
                        continue
            
            if proposal_value:
                return self._format_proposal_value(proposal_value, indexante)
        
        return '-'

    def _format_proposal_value(self, proposal_value, indexante):
        """Format proposal value based on type and indexante"""
        try:
            # Extract numeric value
            if proposal_value.startswith('k ='):
                k_value = float(proposal_value.replace('k =', '').strip().replace(',', '.'))
            elif proposal_value.startswith('k='):
                k_value = float(proposal_value.replace('k=', '').strip().replace(',', '.'))
            elif proposal_value.startswith('1.º Trim. ='):
                k_value = float(proposal_value.replace('1.º Trim. =', '').strip().replace(',', '.'))
            elif proposal_value.startswith('Qualificado com '):
                k_value = float(proposal_value.replace('Qualificado com ', '').strip().replace(',', '.'))
            else:
                k_value = float(proposal_value.replace(',', '.'))
            
            # Convert to €/MWh
            mwh_value = k_value * 1000
            formatted_value = f"{mwh_value:.3f}".replace('.', ',')
            
            # Format based on indexante
            if indexante and indexante in ["TTF (1 -1, 1)", "MibGás (1 -1, 1)"]:
                return f"k={formatted_value}"
            else:
                return formatted_value
                
        except (ValueError, AttributeError) as e:
            print(f"Error formatting proposal value '{proposal_value}': {e}")
            return '-'

    def _write_updated_csv(self, csv_file, headers, rows, price_dates, start_supply_months, 
                          winner_proposals, companies_with_proposals, delimiter):
        """Write updated CSV with new columns"""
        # Prepare new rows
        new_rows = []
        for row, proposal, companies, price_date, start_supply_month in zip(
            rows, winner_proposals, companies_with_proposals, price_dates, start_supply_months):
            
            new_row = row[:29] if len(row) > 29 else row.copy()
            while len(new_row) < 29:
                new_row.append('')
            
            # Calculate profit margin
            profit_margin = '-'
            if proposal and proposal.startswith('k='):
                profit_margin = proposal[2:]
            
            # Add new columns
            new_row.extend([price_date, start_supply_month, proposal, profit_margin, companies])
            
            if len(new_row) > 34:
                new_row = new_row[:34]
            
            new_rows.append(new_row)
        
        # Prepare headers
        base_headers = headers[:29] if len(headers) >= 29 else headers.copy()
        while len(base_headers) < 29:
            base_headers.append(f"Column_{len(base_headers) + 1}")
        
        new_headers = base_headers + ["Price Date", "Start supply month", "Proposta_Vencedor", 
                                     "Profit_Margin (€/MWh)", "Empresas_Com_Propostas"]
        
        if len(new_headers) > 34:
            new_headers = new_headers[:34]
        
        # Write updated CSV
        with open(csv_file, 'w', newline='', encoding='utf-8-sig') as f:
            csv_writer = csv.writer(f, delimiter=delimiter, quotechar='"', quoting=csv.QUOTE_MINIMAL)
            csv_writer.writerow(new_headers)
            for row in new_rows:
                row = ['' if cell == 'None' else str(cell).strip() for cell in row]
                csv_writer.writerow(row)

    def add_proposal_columns(self, csv_file):
        """Add principal proposal and difference columns"""
        # Read CSV
        with open(csv_file, 'r', encoding='utf-8-sig') as f:
            content = f.read(1024)
            delimiter = ';' if ';' in content else ','
            f.seek(0)
            
            reader = csv.reader(f, delimiter=delimiter, quotechar='"')
            headers = next(reader)
            rows = list(reader)
        
        # Find column indices
        col_indices = {}
        for idx, header in enumerate(headers):
            header = header.strip()
            if header == "PROPOSTO PRINCIPAL (€/kWh).PE":
                col_indices['pe'] = idx
            elif header == "PROPOSTO PRINCIPAL (€/kWh).K":
                col_indices['k'] = idx
            elif header == "Proposta_Vencedor":
                col_indices['winner_prop'] = idx
            elif header == "Principal_Proposal":
                col_indices['principal_prop'] = idx
            elif header == "Principal_Difference":
                col_indices['principal_diff'] = idx
            elif header == "VENCEDOR":
                col_indices['winner'] = idx
        
        # Add new headers if needed
        new_headers = headers.copy()
        if 'principal_prop' not in col_indices:
            new_headers.append("Principal_Proposal")
            col_indices['principal_prop'] = len(new_headers) - 1
        
        if 'principal_diff' not in col_indices:
            new_headers.append("Principal_Difference")
            col_indices['principal_diff'] = len(new_headers) - 1
        
        # Process rows
        new_rows = []
        for row in rows:
            new_row = row.copy()
            while len(new_row) <= max(col_indices['principal_prop'], col_indices['principal_diff']):
                new_row.append('')
            
            # Skip if row is too short
            required_idx = max(col_indices['pe'], col_indices['k'], col_indices['winner_prop'], col_indices['winner'])
            if len(row) <= required_idx:
                new_row[col_indices['principal_prop']] = '-'
                new_row[col_indices['principal_diff']] = '-'
                new_rows.append(new_row)
                continue
            
            winner = row[col_indices['winner']].strip()
            winner_prop = row[col_indices['winner_prop']].strip()
            
            # Handle special cases
            if winner.lower() in ["todos excluídos", "todos"] or winner in ['-', ''] and winner_prop in ['-', '']:
                new_row[col_indices['principal_prop']] = '-'
                new_row[col_indices['principal_diff']] = '-'
                new_rows.append(new_row)
                continue
            
            # Calculate principal proposal
            principal_proposal = self._calculate_principal_proposal(row, col_indices, winner_prop)
            new_row[col_indices['principal_prop']] = principal_proposal
            
            # Calculate difference
            principal_diff = self._calculate_proposal_difference(principal_proposal, winner_prop)
            new_row[col_indices['principal_diff']] = principal_diff
            
            new_rows.append(new_row)
        
        # Write updated CSV
        with open(csv_file, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f, delimiter=delimiter, quotechar='"', quoting=csv.QUOTE_MINIMAL)
            writer.writerow(new_headers)
            writer.writerows(new_rows)

    def _calculate_principal_proposal(self, row, col_indices, winner_prop):
        """Calculate the principal proposal value"""
        if winner_prop.startswith('k='):
            value = row[col_indices['k']].strip()
            value_type = 'K'
        else:
            value = row[col_indices['pe']].strip()
            value_type = 'PE'
        
        if value and value not in ['-', 'NAP', 'None', '']:
            try:
                if value.startswith('1.º Trim.'):
                    numeric_part = value.split('=')[1].strip()
                    value_float = float(numeric_part.replace(',', '.'))
                    formatted_value = f"{(value_float * 1000):.3f}".replace('.', ',')
                    return formatted_value
                else:
                    value_float = float(value.replace(',', '.'))
                    formatted_value = f"{(value_float * 1000):.3f}".replace('.', ',')
                    if value_type == 'K':
                        return f"k={formatted_value}"
                    else:
                        return formatted_value
            except (ValueError, AttributeError):
                pass
        
        return '-'

    def _calculate_proposal_difference(self, principal_proposal, winner_proposal):
        """Calculate difference between principal and winner proposals"""
        try:
            if principal_proposal == '-' or winner_proposal == '-':
                return '-'
            
            # Handle k= values
            if principal_proposal.startswith('k=') and winner_proposal.startswith('k='):
                principal_num = float(principal_proposal.replace('k=', '').strip().replace(',', '.'))
                winner_num = float(winner_proposal.replace('k=', '').strip().replace(',', '.'))
                diff = principal_num - winner_num
                return f"k={diff:.3f}".replace('.', ',')
            else:
                # Regular values
                principal_num = float(principal_proposal.replace(',', '.'))
                winner_num = float(winner_proposal.replace(',', '.'))
                diff = principal_num - winner_num
                return f"{diff:.3f}".replace('.', ',')
                
        except (ValueError, AttributeError) as e:
            print(f"Error calculating difference: {e}")
            return '-'

    def process_complete_pipeline(self, excel_path, target_sheet, data_sheet, 
                                 mibgas_excel, output_csv, year=None):
        """Run the complete processing pipeline"""
        print("Step 1: Creating CSV from Excel...")
        self.create_csv_from_excel(excel_path, target_sheet, data_sheet, output_csv, year)
        
        print("Step 2: Extracting MIBGAS data...")
        self.extract_mibgas_data(mibgas_excel, output_csv)
        
        print("Step 3: Adding proposal columns...")
        self.add_proposal_columns(output_csv)
        
        print(f"Pipeline complete! Output saved to {output_csv}")


def main():
    """Main function with argument parsing"""
    parser = argparse.ArgumentParser(description="Generic Gas Tender Processing System")
    parser.add_argument("--excel_path", required=True, help="Path to main Excel file")
    parser.add_argument("--target_sheet", required=True, help="Target sheet name")
    parser.add_argument("--data_sheet", required=True, help="Data sheet name")
    parser.add_argument("--mibgas_excel", required=True, help="Path to MIBGAS Excel file")
    parser.add_argument("--output_csv", required=True, help="Output CSV file path")
    parser.add_argument("--year", type=int, help="Processing year (affects method used)")
    
    args = parser.parse_args()
    
    processor = GasTenderProcessor()
    processor.process_complete_pipeline(
        args.excel_path, args.target_sheet, args.data_sheet,
        args.mibgas_excel, args.output_csv, args.year
    )


if __name__ == "__main__":
    # Run main if called with arguments, otherwise show info
    if len(argparse.sys.argv) > 1:
        main()
    else:
        processor = GasTenderProcessor()
        print("Gas Tender Processing System - Generic Implementation")
        print("Contains all logic from individual scripts with company references removed")
        print("\nUsage:")
        print("python gas_tender_processor.py --excel_path <path> --target_sheet <sheet> --data_sheet <sheet> --mibgas_excel <path> --output_csv <path> [--year <year>]")
        print("\nFeatures:")
        print("- CSV creation from Excel with data cleaning")
        print("- MIBGAS data extraction and merging")
        print("- Proposal calculations and comparisons")
        print("- Price date and supply month calculations")
        print("- Company proposal analysis") 